"""
Table exporter for DocumentAgent pipeline output.

Exports each detected TABLE block to:
  - Individual CSV files  (one per table)
  - A single Excel workbook (one sheet per table)

Usage
-----
    from document_agent.exporters.tables import export_tables

    paths = export_tables(output, output_dir="output/tables")
    # → ["output/tables/p1_b12_revenue.csv", ...]

    xlsx = export_tables_excel(output, "output/tables/all_tables.xlsx")
"""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _safe_filename(text: str, max_len: int = 40) -> str:
    """Convert arbitrary text to a safe filename fragment."""
    slug = re.sub(r"[^\w\s-]", "", text.lower())
    slug = re.sub(r"[\s_-]+", "_", slug).strip("_")
    return slug[:max_len] if slug else "table"


def _table_rows(block: Dict) -> Tuple[List[str], List[Dict]]:
    """Extract (headers, rows) from a TABLE block payload."""
    p = block.get("payload", {})
    headers = p.get("headers") or []
    rows    = p.get("rows")    or []
    return [str(h) for h in headers], rows


def _table_name(block: Dict, idx: int) -> str:
    """Derive a short descriptive name for a table block."""
    p     = block.get("payload", {})
    title = (p.get("title") or "").strip()
    if title:
        return _safe_filename(title)
    summary = (p.get("summary") or "").strip()
    if summary:
        return _safe_filename(summary[:40])
    return f"table_{idx + 1}"


def export_tables(
    output: Dict[str, Any],
    output_dir: str | Path = ".",
    include_metadata: bool = True,
) -> List[Path]:
    """
    Write one CSV file per TABLE block.

    Parameters
    ----------
    output           : DocumentAgent output dict
    output_dir       : directory to write CSVs into
    include_metadata : prepend two comment rows (title, summary) to each CSV

    Returns list of written Paths.
    """
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    blocks = output.get("blocks", [])
    tables = [b for b in blocks if b.get("type") == "table"]
    written: List[Path] = []

    for idx, block in enumerate(tables):
        headers, rows = _table_rows(block)
        if not headers:
            continue

        name     = _table_name(block, idx)
        page     = block.get("page_index", 0) + 1
        bid      = block.get("id", f"b{idx}")
        filename = f"p{page}_{bid}_{name}.csv"
        path     = out_dir / filename

        p = block.get("payload", {})
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if include_metadata:
                title   = (p.get("title")   or "").strip()
                summary = (p.get("summary") or "").strip()
                notes   = (p.get("notes")   or "").strip()
                if title:
                    writer.writerow([f"# {title}"])
                if summary:
                    writer.writerow([f"# {summary}"])
                if notes:
                    writer.writerow([f"# Notes: {notes}"])
                writer.writerow([])  # blank separator

            writer.writerow(headers)
            for row in rows:
                cells = []
                for h in headers:
                    v = row.get(h)
                    cells.append("" if v is None else str(v))
                writer.writerow(cells)

        written.append(path)

    return written


def export_tables_excel(
    output: Dict[str, Any],
    output_path: str | Path,
) -> Optional[Path]:
    """
    Write all TABLE blocks into a single Excel workbook (one sheet per table).

    Requires openpyxl. Returns None if openpyxl is not installed.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    blocks = output.get("blocks", [])
    tables = [b for b in blocks if b.get("type") == "table"]
    if not tables:
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="2563EB")  # blue
    header_align = Alignment(horizontal="center", wrap_text=True)

    for idx, block in enumerate(tables):
        headers, rows = _table_rows(block)
        if not headers:
            continue

        name      = _table_name(block, idx)[:31]  # Excel sheet name max 31 chars
        page      = block.get("page_index", 0) + 1
        sheet_name = f"p{page}_{name}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        p = block.get("payload", {})
        # Metadata rows
        row_offset = 1
        for meta_key in ("title", "summary"):
            val = (p.get(meta_key) or "").strip()
            if val:
                ws.cell(row=row_offset, column=1, value=val).font = Font(italic=True)
                ws.merge_cells(
                    start_row=row_offset, start_column=1,
                    end_row=row_offset,   end_column=max(1, len(headers)),
                )
                row_offset += 1
        if row_offset > 1:
            row_offset += 1  # blank spacer after metadata

        # Header row
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_offset, column=col, value=h)
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = header_align

        # Data rows
        for row in rows:
            row_offset += 1
            for col, h in enumerate(headers, start=1):
                v = row.get(h)
                ws.cell(row=row_offset, column=col, value="" if v is None else str(v))

        # Auto-size columns (approximate)
        for col_idx, h in enumerate(headers, start=1):
            col_values = [str(h)] + [
                str(r.get(h) or "") for r in rows
            ]
            max_len = max((len(v) for v in col_values), default=8)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
